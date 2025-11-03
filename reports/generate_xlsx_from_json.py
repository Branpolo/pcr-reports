#!/usr/bin/env python3
"""
Generate XLSX export from combined JSON report
Exports all error data except curves/readings, pathogen target only
"""

import json
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def extract_target_data(well_curves, well_id):
    """Extract pathogen target data (CT, comments) from well_curves"""
    well_data = well_curves.get(str(well_id)) or well_curves.get(well_id)
    if not well_data:
        return None, None

    targets = well_data.get('targets', [])
    if not targets:
        return None, None

    # Find pathogen target (not IC)
    pathogen_target = None
    if isinstance(targets, list):
        for t in targets:
            if not t.get('is_ic', 0):
                pathogen_target = t
                break
        # If all are IC, take first
        if not pathogen_target and targets:
            pathogen_target = targets[0]
    elif isinstance(targets, dict):
        # Find first non-IC target
        for tname, tdata in targets.items():
            if not tdata.get('is_ic', 0):
                pathogen_target = tdata
                break
        if not pathogen_target and targets:
            pathogen_target = list(targets.values())[0]

    if not pathogen_target:
        return None, None

    ct_value = pathogen_target.get('ct') or pathogen_target.get('machine_ct')

    # Get comments if available
    comments = well_data.get('comments', [])
    comment_text = '; '.join([c.get('text', '') for c in comments[:3]]) if comments else ''

    return ct_value, comment_text


def create_sample_sheet(wb, errors, well_curves, report_name):
    """Create sheet for sample/control reports"""
    ws = wb.create_sheet(report_name)

    # Define columns
    columns = [
        'Sample Name', 'Well Number', 'Mix', 'Run Name', 'Run ID',
        'Extraction Date', 'Error Code', 'Error Message', 'LIMS Status',
        'Clinical Category', 'CT Value', 'Comments'
    ]

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_idx, error in enumerate(errors, 2):
        well_id = error.get('well_id')
        ct_value, comments = extract_target_data(well_curves, well_id)

        row_data = [
            error.get('sample_name', ''),
            error.get('well_number', ''),
            error.get('mix_name', ''),
            error.get('run_name', ''),
            error.get('run_id', ''),
            error.get('extraction_date', ''),
            error.get('error_code', ''),
            error.get('error_message', ''),
            error.get('lims_status', ''),
            error.get('clinical_category', ''),
            ct_value if ct_value else '',
            comments if comments else ''
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def create_discrepancy_sheet(wb, errors, well_curves):
    """Create sheet for discrepancy report"""
    ws = wb.create_sheet('Classification Errors')

    # Define columns
    columns = [
        'Sample Name', 'Sample Name (LIMS)', 'Well Number', 'Mix', 'Run Name', 'Run ID',
        'Extraction Date', 'LIMS Status', 'Machine Classification', 'Final Classification',
        'Error Code', 'Error Message', 'Resolution Codes', 'Clinical Category',
        'Machine CT', 'Target Name', 'Targets Reviewed', 'Comments'
    ]

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_idx, error in enumerate(errors, 2):
        well_id = error.get('well_id')
        ct_value, comments = extract_target_data(well_curves, well_id)

        # Map classifications
        machine_cls = error.get('machine_cls')
        final_cls = error.get('final_cls')
        machine_cls_str = 'POS' if machine_cls == 1 else 'NEG' if machine_cls == 0 else str(machine_cls) if machine_cls is not None else ''
        final_cls_str = 'POS' if final_cls == 1 else 'NEG' if final_cls == 0 else str(final_cls) if final_cls is not None else ''

        # Convert targets_reviewed to string if it's a list/dict
        targets_reviewed = error.get('targets_reviewed', '')
        if isinstance(targets_reviewed, (list, dict)):
            targets_reviewed = str(targets_reviewed)

        row_data = [
            error.get('sample_name', ''),
            error.get('sample_name_lims', ''),
            error.get('well_number', ''),
            error.get('mix_name', ''),
            error.get('run_name', ''),
            error.get('run_id', ''),
            error.get('extraction_date', ''),
            error.get('lims_status', ''),
            machine_cls_str,
            final_cls_str,
            error.get('error_code', ''),
            error.get('error_message', ''),
            error.get('resolution_codes', ''),
            error.get('clinical_category', ''),
            error.get('machine_ct', ct_value) if error.get('machine_ct') else (ct_value if ct_value else ''),
            error.get('target_name', ''),
            targets_reviewed,
            comments if comments else ''
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15


def create_valid_results_sheet(wb, valid_results, error_statistics):
    """Create sheet for valid results summary by mix"""
    ws = wb.create_sheet('Valid Results Summary')

    # Define columns
    columns = [
        'Mix Name',
        'Samples Detected',
        'Samples Not Detected',
        'Controls Passed',
        'Controls Total',
        'SOP Errors',
        'SOP Errors Affected Result',
        'Control Errors',
        'Control Errors Affected Result',
        'Samples Affected by Control Errors',
        'Classification Errors',
        'Classification Errors Affected Result'
    ]

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge valid_results and error_statistics - get all mix names from both
    all_mix_names = set(valid_results.keys()) | set(error_statistics.keys())

    # Filter out run names (they contain .eds_ or .sds_ or are very long)
    # Real mix names are short alphanumeric like "BKV", "CMVQ2", etc.
    real_mix_names = [
        m for m in all_mix_names
        if '.eds_' not in m and '.sds_' not in m and len(m) < 30
    ]

    # Write data - sort by mix name
    sorted_mixes = sorted(real_mix_names)
    for row_idx, mix_name in enumerate(sorted_mixes, 2):
        valid_stats = valid_results.get(mix_name, {})
        error_stats = error_statistics.get(mix_name, {})

        row_data = [
            mix_name,
            valid_stats.get('samples_detected', 0),
            valid_stats.get('samples_not_detected', 0),
            valid_stats.get('controls_passed', 0),
            valid_stats.get('controls_total', 0),
            error_stats.get('sop_errors', 0),
            error_stats.get('sop_errors_affected', 0),
            error_stats.get('control_errors', 0),
            error_stats.get('control_errors_affected', 0),
            error_stats.get('samples_affected_by_controls', 0),
            error_stats.get('classification_errors', 0),
            error_stats.get('classification_errors_affected', 0),
        ]

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def generate_xlsx_from_json(json_file, output_file):
    """Generate XLSX from combined JSON report"""
    # Load JSON
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    reports = data.get('reports', {})

    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create sheets for each report
    if 'sample' in reports:
        sample_data = reports['sample']
        create_sample_sheet(
            wb,
            sample_data.get('errors', []),
            sample_data.get('well_curves', {}),
            'Sample SOP Errors'
        )

    if 'control' in reports:
        control_data = reports['control']
        create_sample_sheet(
            wb,
            control_data.get('errors', []),
            control_data.get('well_curves', {}),
            'Control Errors'
        )

    if 'discrepancy' in reports:
        discrepancy_data = reports['discrepancy']
        create_discrepancy_sheet(
            wb,
            discrepancy_data.get('errors', []),
            discrepancy_data.get('well_curves', {})
        )

    # Create valid results summary sheet
    valid_results = data.get('valid_results', {})
    error_statistics = data.get('error_statistics', {})
    if valid_results or error_statistics:
        create_valid_results_sheet(wb, valid_results, error_statistics)
        mix_count = len(set(valid_results.keys()) | set(error_statistics.keys()))
        print(f"  Added valid results summary: {mix_count} mixes")

    # Save workbook
    wb.save(output_file)
    print(f"XLSX report written to {output_file}")


def main():
    parser = argparse.ArgumentParser(description='Generate XLSX from combined JSON report')
    parser.add_argument('--json', required=True, help='Path to combined JSON file')
    parser.add_argument('--output', help='Output XLSX file path (default: auto-generated)')

    args = parser.parse_args()

    # Generate output filename if not specified
    if args.output:
        output_file = args.output
    else:
        # Use same naming convention as JSON
        if 'combined_report' in args.json:
            output_file = args.json.replace('.json', '.xlsx')
        else:
            output_file = args.json.replace('.json', '_export.xlsx')

    generate_xlsx_from_json(args.json, output_file)


if __name__ == '__main__':
    main()
